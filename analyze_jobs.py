from crawler_upwork import start_crawling
from get_datas_for_bard_results import get_bard_result
import uuid
import os
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border, colors
import re
import time

base_dir = os.getcwd()
cv_desc_path = os.path.join(base_dir,'cv_files','cv.txt')
crawler_datas_save_path = os.path.join(base_dir,'crawler_datas',str(uuid.uuid4().hex) + '.csv')
result_path = os.path.join(base_dir,'results',str(uuid.uuid4().hex) + '.xlsx')


def sayi_olmayanlari_cikar(string):
    yeni_string = ""
    for karakter in string:
        if karakter.isdigit() or karakter=="," or karakter==".":
            if karakter == ",":
                karakter = "."
            yeni_string += karakter
    return yeni_string

def get_percentage(result):
    
    percentage = ""
    
    #index = result.find('%')
    
    percentage_indexs = [i for i, ltr in enumerate(result) if ltr == '%']
    number_indexs = re.findall(r'\d+', result)
    fractional_index = result.find("0.")
    
    try:
        if percentage_indexs:
            for index in percentage_indexs:
                if not float(sayi_olmayanlari_cikar(result[index-9:index])) >= 100:
                    percentage += sayi_olmayanlari_cikar(result[index-9:index]) + 'percentage    '       
        
        elif fractional_index != -1:
            fractional_number = float(sayi_olmayanlari_cikar(result[fractional_index:fractional_index+4]))
            if not fractional_number >= 100:
                percentage += str(fractional_number*100) + "fractional "
        
        elif number_indexs:
            for index in percentage_indexs:
                percentage += index + 'digit    '
    except:
        print("Error",result)
    return percentage
    
    
def get_datas_sorted_by_matching(all_datas):
    result_limit = 10
    results = sorted(all_datas, key=lambda d: d['matching_score'], reverse=True)[0:result_limit]
    return results

def create_result_xlsx(all_datas,result_path):

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'İş Sonuçları'
    
    columns = ['Id',
               'Job Title',
               'Job Type',
               'Job Description',
               'Bard Result',
               'Matching']
    
    row_num = 1
    
    #Adding titles
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
    
    
    #Adding Datas
    for data in all_datas:
        row = []
        row_num += 1
        row.append(data["id"])
        row.append(data["title"])
        row.append(data["type"])
        row.append(data["text"])
        row.append(data["bard_result"])
        row.append(data["matching_score"])
     
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
        
     
    
    workbook.save(result_path)

    print("Tamamlandı...")


def main():
    print('Enter keyword: ')
    keyword = input()
    all_datas = start_crawling(keyword,crawler_datas_save_path)
    cv_desc = open(cv_desc_path, "r").read()
    
    
    print('*'*100)
    print(cv_desc)
    print('*'*100)
    
    for index,crawl_data in enumerate(all_datas):
        bard_result = get_bard_result(cv_desc,crawl_data)
        crawl_data['bard_result'] = bard_result
        crawl_data['matching_score'] = get_percentage(bard_result)
        print("Bard Sonucu Getiriliyor",index)
        time.sleep(2)
        
    #sorted_results = get_datas_sorted_by_matching(all_datas)
    
    create_result_xlsx(all_datas,result_path)

main()